#!/usr/bin/env python3
"""Build the design matrix for flow cytometry panel design."""

import os
import sys
import argparse
import logging
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from tools.file_io import read_excel_or_csv
from tools.list_tools import items_in_a_not_b
from tools.df_tools import (fillna, reset_index, append_dataframe,
                            dataframe_row_from_named_list, stranspose)
from functions.preprocessing import (preprocess_instrument_config,
                                     preprocess_antibody_inventory)


def main():
    parser = argparse.ArgumentParser(description='Build design matrix')
    parser.add_argument('-i', '--input-file', default='data/design/panel.csv',
                        help='Panel input file')
    parser.add_argument('-a', '--antibody-inventory', default='ref/antibody_inventory.xlsx',
                        help='Antibody inventory file')
    parser.add_argument('-c', '--instrument-config', default='ref/instrument_config.xlsx',
                        help='Instrument configuration file')
    parser.add_argument('-o', '--output-dir', default='data/design',
                        help='Output directory')
    parser.add_argument('-t', '--troubleshooting', action='store_true')

    args = parser.parse_args()

    # Setup
    wd = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_dir = os.path.join(wd, args.output_dir)

    # Logging
    start_time = datetime.now()
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
    logging.info(f'Script started at: {start_time}')

    # Load configuration files
    instr_cfg = read_excel_or_csv(os.path.join(wd, args.instrument_config))
    instr_cfg = preprocess_instrument_config(instr_cfg)

    # Expand fluorophores to long format
    instr_cfg_long = instr_cfg.copy()
    if 'fluorophore' in instr_cfg_long.columns:
        instr_cfg_long = instr_cfg_long.assign(
            fluorophore=instr_cfg_long['fluorophore'].str.split(', ')
        ).explode('fluorophore')

    # Load antibody inventory
    ab_inv = read_excel_or_csv(os.path.join(wd, args.antibody_inventory))
    ab_inv = preprocess_antibody_inventory(ab_inv)

    # Read panel
    panel = read_excel_or_csv(os.path.join(wd, args.input_file))

    # Match antibodies
    ab_shortlist = panel[['antibody']].merge(
        ab_inv, on='antibody', how='inner'
    )

    antibodies_not_found = sorted(
        set(panel['antibody'].unique()) - set(ab_inv['antibody'].unique())
    )

    # Save shortlist
    if not args.troubleshooting:
        os.makedirs(output_dir, exist_ok=True)
        basename = os.path.splitext(os.path.basename(args.antibody_inventory))[0]
        filepath = os.path.join(output_dir, f'{basename}-shortlist.csv')
        ab_shortlist.to_csv(filepath, index=False)

        if antibodies_not_found:
            troubleshooting_dir = os.path.join(output_dir, 'troubleshooting')
            os.makedirs(troubleshooting_dir, exist_ok=True)
            with open(os.path.join(troubleshooting_dir, 'antibodies_not_found.txt'), 'w') as f:
                f.write('\n'.join(antibodies_not_found))

    # Build base table
    ab_counts = ab_shortlist[['antibody', 'company', 'catalog_no', 'fluorophore']].merge(
        instr_cfg_long[['fluorophore', 'channel_id']],
        on='fluorophore', how='left'
    )
    ab_counts['channel_id'] = ab_counts['channel_id'].fillna('other')

    # Group by antibody and channel
    ab_counts = ab_counts.groupby(['antibody', 'channel_id']).agg(
        num_fluorophores=('fluorophore', 'nunique'),
        fluorophores=('fluorophore', lambda x: ', '.join(x.unique())),
        most_common_fluorophore=('fluorophore', lambda x: x.value_counts().index[0])
    ).reset_index()

    # Pivot to design matrix
    design_matrix = ab_counts.pivot_table(
        index='antibody',
        columns='channel_id',
        values='num_fluorophores',
        fill_value=0
    ).reset_index()

    # Sort rows by number of channels
    numeric_cols = design_matrix.select_dtypes(include=[np.number]).columns
    design_matrix['num_channels_per_ab'] = (design_matrix[numeric_cols] >= 1).sum(axis=1)
    design_matrix = design_matrix.sort_values('num_channels_per_ab')

    # Sort columns by emission
    if 'bandpass_filter' in instr_cfg.columns:
        channel_order = (instr_cfg
                         .sort_values(['bandpass_filter', 'excitation'],
                                      ascending=[True, False])
                         ['channel_id'].tolist())

        existing_cols = [c for c in channel_order if c in design_matrix.columns]
        other_cols = items_in_a_not_b(design_matrix.columns.tolist(),
                                      ['antibody', 'num_channels_per_ab'] + channel_order)
        design_matrix = design_matrix[
            ['antibody', 'num_channels_per_ab'] + existing_cols + other_cols
        ]

    # Save to Excel
    if not args.troubleshooting:
        wb = Workbook()
        ws = wb.active

        # Write data
        for r_idx, row in enumerate(design_matrix.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx + 1, column=c_idx, value=value)

        # Write headers
        for c_idx, col in enumerate(design_matrix.columns, 1):
            ws.cell(row=1, column=c_idx, value=col)

        # Add highlighting for non-zero values
        gray_fill = PatternFill(start_color='C3C3C3', end_color='C3C3C3',
                                fill_type='solid')

        for r_idx, row in enumerate(design_matrix.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (int, float)) and value > 0:
                    ws.cell(row=r_idx, column=c_idx).fill = gray_fill

        filepath = os.path.join(output_dir, 'design_matrix.xlsx')
        wb.save(filepath)

    end_time = datetime.now()
    logging.info(f'Script ended at: {end_time}')
    logging.info(f'Script completed in: {end_time - start_time}')


if __name__ == '__main__':
    main()
